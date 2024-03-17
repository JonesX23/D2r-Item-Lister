import os
import glob
import pytesseract
import cv2
import openpyxl
from datetime import datetime
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
# Funktion zum Extrahieren des Textes aus dem Bild und Schreiben in eine Excel-Datei
def extract_text_and_write_to_excel(image_path, output_excel):
    # Bild einlesen
    image = cv2.imread(image_path)

    # Text aus dem Bild extrahieren
    text = pytesseract.image_to_string(image).strip()

    # Excel-Datei öffnen oder erstellen
    if os.path.exists(output_excel):
        wb = openpyxl.load_workbook(output_excel)
        ws = wb.active
        # Suche nach der nächsten leeren Zeile
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        next_row = 1

    # Prüfen, ob "charm" im Text enthalten ist
    if "charm" in text.lower():
        # Wenn "charm" gefunden wird, schreibe die erste Zeile und alle folgenden Zeilen, außer Zeile 2 und 3
        lines = text.split('\n')
        for i, line in enumerate(lines):
            if i == 0:
                ws.cell(row=next_row, column=2).value = line
            elif i not in (1, 2):
                ws.cell(row=next_row, column=2).value += f"\n{line}"
    else:
        # Wenn "charm" nicht gefunden wird, schreibe den gesamten Text in die erste Spalte
        ws.cell(row=next_row, column=2).value = text

    # Schreibe den Timestamp in die zweite Spalte
    #timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
   # ws.cell(row=next_row, column=3).value = timestamp

    # Schreibe die Nummerierung in die erste Spalte
    ws.cell(row=next_row, column=1).value = next_row

    # Excel-Datei speichern
    wb.save(output_excel)

# Pfad zum Ordner mit den Bildern
image_folder = r"C:\Users\jonas\Pictures\Screenshots"

# Suchen nach allen Bilddateien im Ordner, die heute erstellt wurden
today = datetime.today().date()
list_of_images = glob.glob(os.path.join(image_folder, '*.png')) + glob.glob(os.path.join(image_folder, '*.jpg'))
today_images = [img for img in list_of_images if datetime.fromtimestamp(os.path.getctime(img)).date() == today]

# Ausgabedatei für Excel
output_excel = r"C:\Users\jonas\Desktop\LOOT.xlsx"

# Extrahiere den Text aus allen heutigen Bildern und schreibe sie in die Excel-Datei
for image_path in today_images:
    extract_text_and_write_to_excel(image_path, output_excel)
